import fasttext
import openpyxl
from utility import get_semantic_noun_neigbours, get_unique_prefixes, is_unique, most_frequent
from gensim.models.fasttext import load_facebook_model
from gensim.similarities.annoy import AnnoyIndexer

'''
Joans model replication, unique prefixing -> semantic neighbours -> classifer
'''

workbook = openpyxl.load_workbook(f'./DataFiles/ZuluNounsSingleClass.xlsx')
worksheet = workbook.active

# Get all the class labels from the second column
nouns = [row[0].lower() for row in worksheet.iter_rows(min_row=1, values_only=True)]

workbook = openpyxl.load_workbook(f'./DataFiles/ZuluNouns20%TestSet.xlsx')
worksheet = workbook.active

# Get all the class labels from the second column
test_nouns_and_classes = [(row[0].lower(),row[1]) for row in worksheet.iter_rows(min_row=1, values_only=True)]


workbook.close()

prefix_dictionary = [
    ('1',"um"),
    ('1',"umu"),
    ('1a',"u"),
    ('2',"aba"),
    ('2',"abe"),
    ('2a',"o"),
    ('3',"umu"),
    ('3',"um"),
    ('4',"imi"),
    ('5',"i"),
    ('5',"ili"),
    ('6',"ama"),
    ('6',"ame"),
    ('7',"is"),
    ('7',"isi"),
    ('8',"iz"),
    ('8',"izi"),
    ('9',"in"),
    ('9',"im"),
    ('10',"izin"),
    ('10',"izim"),
    ('11',"u"),
    ('11',"ulu"),
    ('14',"ubu"),
    ('14',"ub"),
    ('15',"uku"),
    ('15',"ukw"),
]

#load the semantic model
semantic_model = load_facebook_model('./ModelsAndVectors/zu_fasttext_embeddings.bin')

#annoy index
annoyIndex = AnnoyIndexer(model=semantic_model, num_trees=10)

#load the classifer model
classifier_model = fasttext.load_model('./ModelsAndVectors/zu_fasttext_classifer.bin')

#get our the unique prefixies from our list
unique_prefixes_classes = get_unique_prefixes(prefixDict=prefix_dictionary)

correct=0
guesses=0


for query_word, correct_class in test_nouns_and_classes:
    
    guesses+=1
    print(str(guesses), end='\r')

    value = is_unique(unique_prefixes_classes, query=query_word)
    
    #Check for unique prefix
    if(value):
        if correct_class==value:
            correct+=1
    else:
        # get semantic neigbours
        semantic_similar_words = get_semantic_noun_neigbours(nouns=nouns,semantic_model=semantic_model,indexer=annoyIndex,query_word=query_word,topn=40)

        predicted_labels = []
        
        #go through all neigbours and predict conocrd and class labels
        for word in semantic_similar_words:
            prediction = classifier_model.predict(word, k=20)

            labels, arr = prediction

            labels = list(labels)
            
            noun_labels = filter((lambda x: "SC" not in x),labels)
            concord_labels = filter((lambda x: "SC" in x),labels)


            #filter predictions based on agreement with prediction of concord and class
            final_noun_class = list(noun_labels)[0].replace("__label__","")
            final_concord_class = list(concord_labels)[0].replace("__label__SC","")

            if final_noun_class == final_concord_class:
                predicted_labels.append(final_concord_class)
            
        #make a final prediction based on frequency
        predicted_label = most_frequent(predicted_labels)
        if correct_class == predicted_label:
            correct += 1

print(guesses)
print(correct)
print(str(correct/guesses))